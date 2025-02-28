from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter

def find_headers_and_values(file_path, sheet_name, cell_address):
    """
    Finds the row header, column header, and their corresponding values,
    as well as the value of the target cell in an Excel sheet.

    Args:
        file_path: Path to the .xlsx file.
        sheet_name: The name of the sheet.
        cell_address: The target cell address (e.g., "B2").

    Returns:
        A dictionary containing:
        - "row_header_address": The cell address of the row header (e.g., "A2").
        - "row_header_value": The value of the row header cell.
        - "col_header_address": The cell address of the column header (e.g., "B1").
        - "col_header_value": The value of the column header cell.
        - "target_cell_value": The value of the target cell.
        Or None if an error occurs.
    """
    try:
        workbook = load_workbook(file_path, read_only=True)  # Open in read-only mode for efficiency
        sheet = workbook[sheet_name]

        # Parse the cell address to get row and column numbers
        col_num, row_num = coordinate_to_tuple(cell_address)

        # Find the row header (same row, first column)
        row_header_address = f"A{row_num}"
        row_header_value = sheet[row_header_address].value

        # Find the column header (first row, same column)
        col_header_address = f"{get_column_letter(col_num)}1"
        col_header_value = sheet[col_header_address].value
        
        # Get the target cell value
        target_cell_value = sheet[cell_address].value

        return {
            "row_header_address": row_header_address,
            "row_header_value": row_header_value,
            "col_header_address": col_header_address,
            "col_header_value": col_header_value,
            "target_cell_value": target_cell_value,
        }

    except FileNotFoundError:
        print(f"Error: File not found at '{file_path}'")
        return None
    except KeyError:
        print(f"Error: Sheet '{sheet_name}' not found in the file.")
        return None
    except ValueError:
        print(f"Error: Invalid cell address '{cell_address}'.")
        return None
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return None

def main():
    file_path = "data.xlsx"  # Replace with your file path
    sheet_name = "Time"  # Replace with the sheet name
    cell_address = "B2"  # Replace with your cell address

    result = find_headers_and_values(file_path, sheet_name, cell_address)

    if result:
        print(f"Details for cell {cell_address}:")
        print(f"  Origin: {result['row_header_address']} (Value: {result['row_header_value']})")
        print(f"  Destination: {result['col_header_address']} (Value: {result['col_header_value']})")
        print(f"  Time {cell_address}: (Value: {result['target_cell_value']})")


if __name__ == "__main__":
    main()
