"""
File: write_xlsx.py
Purpose: This program provides a function for writing data to a new sheet 
         within an existing Excel file. It is specifically designed to 
         write data from a dictionary into cell addresses in the sheet.
         Called by calculate_time.py to save the time results in data.xlsx
Author: marcel the coolest
Dependencies:
    - openpyxl
    https://openpyxl.readthedocs.io/en/stable/index.html
"""
from openpyxl import load_workbook


def write_to_xlsx_sheet(data_dict, directory, sheet_name="Sheet1"):
    """
    Writes data from a dictionary to a specified sheet in an existing Excel file.

    Args:
        data_dict: A dictionary where keys are cell addresses (e.g., 'A1', 'B2') and
                   values are data to be written to those cells.
        directory: The path to the existing Excel file.
        sheet_name: The name of the sheet to write to. If it doesn't exist, a new sheet is created.
    """
    try:
        workbook = load_workbook(directory)

        # Check if the sheet exists; create it if it doesn't
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        sheet = workbook[sheet_name]

        # Write data to the sheet
        for cell_address, value in data_dict.items():
            sheet[cell_address] = value[0]  # Assuming value is a tuple, take the first element

        workbook.save(directory)
        print(f"Data successfully written to sheet '{sheet_name}' in '{directory}'.")

    except FileNotFoundError:
        print(f"Error: File '{directory}' not found.")
    except Exception as e:
        print(f"An error occurred while writing to the Excel file: {e}")
