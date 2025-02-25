"""
File: calculate_time.py
Purpose: This program calculates estimated travel times based on Naismith's Rule, 
         taking into account distance, elevation gain, and elevation loss. 
         It reads data from an Excel file, processes it, and writes the 
         calculated times back to the Excel file (data.xlsx). The results should be 
         used as weights in Dijkstra's Algorithm. 
Author: marcel the coolest
Dependencies:
    - read_xlsx.py
    - write_xlsx.py
    - data.xlsx
    - openpyxl
    https://openpyxl.readthedocs.io/en/stable/index.html
"""

import os
from openpyxl import load_workbook  # Import the load_workbook function from the openpyxl library for working with Excel files.

# Check if 'read_xlsx.py' is in the correct location and import if true
if os.path.exists(os.path.join(os.path.dirname(__file__), 'read_xlsx.py')):
    from read_xlsx import extract_data_from_excel  # Import the function to read data from the Excel file.
    from write_xlsx import write_to_xlsx_sheet # Import the function to write data to the Excel file.
else:
    print("Error: 'read_xlsx.py' or 'write_xlsx.py' not found in the same directory.")
    exit() 

def calculate_from_cells(gain, loss, distance):
    """Calculates time based on Naismith's Rule.
    Total Time (minutes) = ( (distance / 1000) / 5 + (elevation gain / 600) - ( (elevation loss / 300) * 10 / 60 ) ) * 60

    Explanation:
        It converts distance (meters to km) and divides by speed (5 km/h) for travel time.  
        It divides elevation gain by 600 m/h for gain time. 
        It divides elevation loss by 300 m/10min, multiplies by 10 to get minutes, and divides by 60 for hours.  
        The travel and gain times are summed, loss time is subtracted, 
        and the result (in hours) is multiplied by 60 for minutes.

    Args:
        gain: Elevation gain in meters.
        loss: Elevation loss in meters.
        distance: Distance in meters.

    Returns:
        Total time in seconds.
    """
    # Convert distance to km. Caltopo exports in meters. Naismith's uses km
    km = distance / 1000.0
    flat_speed = 5  # km/hour, the speed when it is not up or down hill.
    # Elevation remains in meters. 
    gain_rate = 600  # meters/hour, the rate that the elevation gain is handled.
    loss_rate = 300  # meters per 10 minutes, the rate that elevation loss is handled.

    # Calculate time for each component
    distance_time = km / flat_speed  # hours, the time it will take to travel the distance.
    gain_time = gain / gain_rate  # hours, the time it will take to travel up the elevation gain.
    loss_time = (loss / loss_rate * 10) / 60  # Convert loss time to hours, the time that will be reduced by going down.
    total_time_seconds = (distance_time + gain_time - loss_time) * 60 #the formula to calculate total time.
    return total_time_seconds


def process_data_and_calculate_naismith(data_dict):
    """Processes the dictionary data and calculates Naismith's Rule for each cell.

    Args:
        data_dict: A dictionary where keys are cell addresses and values are tuples
                   in the format (gain, loss, distance).

    Returns:
        A dictionary where keys are cell addresses and values are the calculated times
        in seconds based on Naismith's Rule.
    """
    results = {} # create an empty dictionary to store the results
    for cell_address, (gain, loss, distance) in data_dict.items(): # loops through the keys (cell address) and values (gain, loss, distance)
        try: 
            calculated_time = calculate_from_cells(gain, loss, distance) #use the formula in calculate from cells to get the time
            results[cell_address] = (calculated_time,)  # added comma to make it a single item tuple for write_xlsx.py
        except Exception as e: #if there is an exception, print the following:
            print(f"An error has occurred on cell {cell_address}: {e}")
    return results

def extract_data_from_excel_ignore_extra_sheet(directory):
    """
    Explanation:
        Extracts data from the first three sheets of an Excel file, ignoring
        any additional sheets, and returns a dictionary where keys are cell
        addresses and values are tuples representing (value_sheet1, value_sheet2, value_sheet3).

    Args:
        directory: The path to the Excel file.

    Returns:
        A dictionary with cell addresses as keys and tuples of three values as values.
        Returns None if the file is not found or if an error occurs.
    """
    try: #try the code below
        workbook = load_workbook(directory) #open the workbook
        all_data = {} #create an empty dictionary to store data
        sheet_names = workbook.sheetnames #gets all of the sheet names

        # Use only the first three sheets
        num_sheets_to_use = min(3, len(sheet_names)) #get the smaller number of either 3 or how many sheets are in the excel file
        sheets_to_use = sheet_names[:num_sheets_to_use] #create a list of sheet names, starting from the first, to the smaller number

        # Iterate through rows and cells in the sheets
        for row_num in range(2, workbook[sheets_to_use[0]].max_row + 1): # get the rows, starting at row 2, and going to the last row.
            for col_num in range(2, workbook[sheets_to_use[0]].max_column + 1): #get the columns, starting at column 2, and going to the last column.
                cell_address = workbook[sheets_to_use[0]].cell(row=row_num, column=col_num).coordinate #get the cell address
                data_tuple = [] #create an empty list to hold the cell values.

                # Collect the value of that cell address from each sheet
                for sheet_name in sheets_to_use: # go through each of the sheets in sheets_to_use
                    sheet = workbook[sheet_name] #get the sheet
                    cell_value = sheet[cell_address].value #get the value from the cell in that sheet
                    data_tuple.append(cell_value) #add the value to the list.

                # Add this data to all_data, if all the sheets are not none
                if all(value is not None for value in data_tuple): #if all of the values in data_tuple are not none.
                    all_data[cell_address] = tuple(data_tuple) #create a tuple and add it to the dictionary.


        return all_data # return the dictionary


    except FileNotFoundError:
        print(f"Error: File '{directory}' not found.")
        return None  
    except ValueError as ve: 
      print(f"Error: {ve}")
      return None
    except Exception as e: # if any other type of error is found
        print(f"An error occurred: {e}")
        return None



def main():
    """
    Main function to orchestrate the data extraction, Naismith's rule calculation.
    """
    # Use the function to get data from the first 3 sheets of Excel
    filename = 'data.xlsx'  # Changed filename here
    script_dir = os.path.dirname(__file__) #gets the current directory that the script is located in.
    directory = os.path.join(script_dir, filename) # adds the filename to the directory to get the full path.

    pulled_data = extract_data_from_excel_ignore_extra_sheet(directory) # gets the data from the excel sheet.

    if pulled_data: # if pulled_data has data in it.
        print("Data extracted from Excel:")
        print(pulled_data)

        # Process the data to get the correct structure
        naismith_ready_data = {} #create an empty dictionary
        for cell, values in pulled_data.items(): # loop through the keys (cell) and values in pulled_data
            # Assuming gain is values[0], loss is values[1], distance is values[2]
            naismith_ready_data[cell] = (values[0], values[1], values[2]) #create the correct tuples for the data.

        # Process the data and calculate Naismith's Rule
        naismith_results = process_data_and_calculate_naismith(naismith_ready_data) # calculate the times

        print("\nNaismith's Rule Results:")
        print(naismith_results)

        # Prepare data for writing to excel
        if naismith_results: #if the results are there.
            filename = 'data.xlsx' # Changed filename here
            script_dir = os.path.dirname(__file__) #gets the directory of the file.
            directory = os.path.join(script_dir, filename) #add the directory and filename together to get the full path.

            write_to_xlsx_sheet(naismith_results,directory,"Time") #write the time results to the excel sheet.

    else: #if pulled data is empty
        print("Failed to extract data from the Excel file.")

# This code makes sure that the 'main' function only runs when you execute this script directly.
# This allows the script to be imported as a module without running 'main', preventing
# errors and allowing the reuse of functions. Without this, importing the file would run the whole program.
if __name__ == "__main__":
    main() #calls the main function
