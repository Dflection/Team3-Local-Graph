"""
File: read_xlsx.py
Purpose: This program provides functions for reading data from Excel files.
         It is designed to extract specific data from multiple sheets and 
         organize it into a dictionary for further processing by calculate_time.
Author: marcel the coolest
Dependencies:
    - openpyxl
    https://openpyxl.readthedocs.io/en/stable/index.html
"""
from openpyxl import load_workbook

def extract_data_from_excel(directory, num_sheets=3):
    """
    Extracts data from an Excel file, reading from a specified number of sheets.

    Args:
      directory: The path to the Excel file to read from (e.g., 'data.xlsx' or 'C:/my_files/data.xlsx').
        - Type: str
        - Required: Yes
      num_sheets: The number of sheets to read from the Excel file, starting from the first sheet.
        - Type: int
        - Required: No, defaults to 3
    Returns:
      A dictionary with cell addresses as keys and tuples of values as values, or None if there was an error.
    """
    try:
        workbook = load_workbook(directory)
        all_data = {}
        sheet_names = workbook.sheetnames
        
        # Use only the specified number of sheets
        num_sheets_to_use = min(num_sheets, len(sheet_names))
        sheets_to_use = sheet_names[:num_sheets_to_use]

        # Iterate through rows and cells in the sheets
        for row_num in range(2, workbook[sheets_to_use[0]].max_row + 1):
            for col_num in range(2, workbook[sheets_to_use[0]].max_column + 1):
                cell_address = workbook[sheets_to_use[0]].cell(row=row_num, column=col_num).coordinate
                data_tuple = []

                # Collect the value of that cell address from each sheet
                for sheet_name in sheets_to_use:
                    sheet = workbook[sheet_name]
                    cell_value = sheet[cell_address].value
                    data_tuple.append(cell_value)

                # Add this data to all_data, if all the sheets are not none
                if all(value is not None for value in data_tuple):
                    all_data[cell_address] = tuple(data_tuple)


        return all_data

    except FileNotFoundError:
        print(f"Error: File '{directory}' not found.")
        return None
    except ValueError as ve:
      print(f"Error: {ve}")
      return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None
